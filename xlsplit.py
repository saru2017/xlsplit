import openpyxl
import datetime
import sys
import os


if len(sys.argv) != 2:
    print("usage: python %s [filename]" % (sys.argv[0]))
    exit(-1)

filename = sys.argv[1]

wb = openpyxl.load_workbook(filename)
sheets_names = wb.sheetnames
wb.close()

for i in range(len(sheets_names)):
    wb = openpyxl.load_workbook(filename)
    remain_name = sheets_names[i]
    for j in range(len(sheets_names)):
        if i == j:
            continue

        wb.remove(wb[sheets_names[j]])
        
    str = os.path.splitext(filename)[0] + "_" + remain_name + ".xlsx"
    wb.active = 1
    wb.save(str)
    wb.close()

